"""
Excel writer for HD Adjustment invoices.
Appends to an existing .xlsx file (or creates one from scratch).
Matches the color formatting of the reference HD_Adjustment_812_All_Invoices.xlsx.
"""

import os
import re
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pdf_parser import HEADER_COLS, LI_COLS
from .base import Writer

# ── fill colors (from the corrected HD_Adjustment_812 template) ───────────────
_FILL_HDR1_INV  = PatternFill("solid", fgColor="FF1F2D3D")  # row-1 group label, invoice section
_FILL_HDR1_LI1  = PatternFill("solid", fgColor="FF1A4A6E")  # row-1 group label, LI1
_FILL_HDR1_LID  = PatternFill("solid", fgColor="FF5D2510")  # row-1 group label, LI2+
_FILL_HDR2_INV  = PatternFill("solid", fgColor="FF344D6E")  # row-2 field header, invoice cols
_FILL_HDR2_LI1  = PatternFill("solid", fgColor="FF1A4A6E")  # row-2 field header, LI1
_FILL_HDR2_LID  = PatternFill("solid", fgColor="FF5D2510")  # row-2 field header, LI2+
_FILL_DATA_INV  = PatternFill("solid", fgColor="FFF9FAFB")  # data rows, invoice cols
_FILL_DATA_LI1  = PatternFill("solid", fgColor="FFEAF4FB")  # data rows, LI1
_FILL_DATA_LID  = PatternFill("solid", fgColor="FFFDF2E9")  # data rows, LI2+

_FONT_HDR  = Font(name='Calibri', size=10, bold=True, color="FFFFFFFF")
_FONT_DATA = Font(name='Calibri', size=10)
_ALIGN_L   = Alignment(horizontal='left', vertical='center', wrap_text=False)
_ALIGN_C   = Alignment(horizontal='center', vertical='center')

# 12 invoice columns + 9 per line item group
_N_INV  = len(HEADER_COLS)   # 12
_N_LI   = len(LI_COLS)       # 9

# Column widths
_INV_WIDTHS = [14, 16, 14, 14, 14, 14, 14, 14, 18, 10, 22, 8]
_LI_WIDTHS  = [12, 10, 18, 12, 12, 7, 7, 11, 11]


def _li_start(li_index: int) -> int:
    """1-based column of the first cell in line-item group li_index (0=LI1, 1=LI2…)."""
    return _N_INV + 1 + li_index * _N_LI


def _write(ws, row: int, col: int, value, fill: PatternFill, font: Font, align: Alignment):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.fill = fill
    cell.font = font
    cell.alignment = align


def _apply_header_row1(ws, max_li: int):
    """Write row 1: merged group labels."""
    # Invoice Details
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_N_INV)
    cell = ws.cell(1, 1)
    cell.value = "Invoice Details"
    cell.fill = _FILL_HDR1_INV
    cell.font = _FONT_HDR
    cell.alignment = _ALIGN_C
    for col in range(2, _N_INV + 1):
        ws.cell(1, col).fill = _FILL_HDR1_INV

    # LI1 and LI2+
    for li in range(max_li + 1):
        s = _li_start(li)
        label = "LINE ITEM 1 · Credit" if li == 0 else f"LINE ITEM {li + 1} · Debit"
        fill = _FILL_HDR1_LI1 if li == 0 else _FILL_HDR1_LID
        ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=s + _N_LI - 1)
        cell = ws.cell(1, s)
        cell.value = label
        cell.fill = fill
        cell.font = _FONT_HDR
        cell.alignment = _ALIGN_C
        for col in range(s + 1, s + _N_LI):
            ws.cell(1, col).fill = fill


def _apply_header_row2(ws, max_li: int):
    """Write row 2: field name headers."""
    for ci, name in enumerate(HEADER_COLS, start=1):
        _write(ws, 2, ci, name, _FILL_HDR2_INV, _FONT_HDR, _ALIGN_L)

    for li in range(max_li + 1):
        s = _li_start(li)
        fill = _FILL_HDR2_LI1 if li == 0 else _FILL_HDR2_LID
        for ci, name in enumerate(LI_COLS, start=0):
            _write(ws, 2, s + ci, name, fill, _FONT_HDR, _ALIGN_L)


def _set_column_widths(ws, max_li: int):
    for ci, w in enumerate(_INV_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for li in range(max_li + 1):
        s = _li_start(li)
        for ci, w in enumerate(_LI_WIDTHS):
            ws.column_dimensions[get_column_letter(s + ci)].width = w


def _current_max_li(ws) -> int:
    """Count how many LI groups exist by scanning row-1 merged cell labels."""
    count = 0
    for merged in ws.merged_cells.ranges:
        if merged.min_row == 1 and merged.min_col > _N_INV:
            count += 1
    return max(count - 1, 0)  # subtract LI1; min 0


class ExcelWriter(Writer):
    """Appends invoices to a .xlsx file, initializing headers on first use."""

    def __init__(self, file_path: str):
        self._path = file_path
        self._wb = None
        self._ws = None

    def _load(self):
        if self._wb is not None:
            return
        if os.path.exists(self._path):
            self._wb = load_workbook(self._path)
            self._ws = self._wb.active
        else:
            self._wb = Workbook()
            self._ws = self._wb.active
            self._ws.title = "Adjustments"

    def _save(self):
        self._wb.save(self._path)

    def is_initialized(self) -> bool:
        if not os.path.exists(self._path):
            return False
        try:
            wb = load_workbook(self._path, read_only=True)
            ws = wb.active
            # Initialized if row 2 has HEADER_COLS[0] in cell A2
            val = ws.cell(2, 1).value
            wb.close()
            return val == HEADER_COLS[0]
        except Exception:
            return False

    def initialize_headers(self, max_line_items: int = 0) -> None:
        self._load()
        _apply_header_row1(self._ws, max_line_items)
        _apply_header_row2(self._ws, max_line_items)
        _set_column_widths(self._ws, max_line_items)
        self._ws.row_dimensions[1].height = 24
        self._ws.row_dimensions[2].height = 20
        self._ws.freeze_panes = 'A3'
        self._save()

    def find_duplicate(self, invoice_num: str) -> bool:
        self._load()
        for row in self._ws.iter_rows(min_row=3, max_col=1, values_only=True):
            if row[0] == invoice_num:
                return True
        return False

    def expand_columns_if_needed(self, num_debit_items: int) -> None:
        """Expand headers if this invoice needs more columns than currently exist."""
        self._load()
        current = _current_max_li(self._ws)
        needed = num_debit_items  # LI2..LI(n+1) for n debit items → index 1..n

        if needed <= current:
            return

        # Clear existing row-1 merges beyond invoice cols
        to_remove = [r for r in self._ws.merged_cells.ranges if r.min_row == 1 and r.min_col > _N_INV]
        for r in to_remove:
            self._ws.merged_cells.remove(r)

        _apply_header_row1(self._ws, needed)
        _apply_header_row2(self._ws, needed)
        _set_column_widths(self._ws, needed)
        self._save()

    def append_invoice(self, invoice: dict) -> None:
        self._load()

        if not self.is_initialized():
            self.initialize_headers(len(invoice.get('debit_items', [])))
        else:
            self.expand_columns_if_needed(len(invoice.get('debit_items', [])))

        # Find next empty row
        next_row = self._ws.max_row + 1
        if next_row < 3:
            next_row = 3

        # Invoice header (cols 1–12)
        inv_cells = self.invoice_header_row(invoice)
        for ci, val in enumerate(inv_cells, start=1):
            _write(self._ws, next_row, ci, val, _FILL_DATA_INV, _FONT_DATA, _ALIGN_L)

        # LI1 — credit summary (cols 13–21)
        li1_cells = self.li1_row(invoice.get('credit_line'))
        s = _li_start(0)
        for ci, val in enumerate(li1_cells):
            _write(self._ws, next_row, s + ci, val, _FILL_DATA_LI1, _FONT_DATA, _ALIGN_L)

        # LI2+ — debit items
        for di, item in enumerate(invoice.get('debit_items', [])):
            s = _li_start(di + 1)
            for ci, val in enumerate(self.debit_row(item)):
                _write(self._ws, next_row, s + ci, val, _FILL_DATA_LID, _FONT_DATA, _ALIGN_L)

        self._ws.row_dimensions[next_row].height = 18
        self._save()
