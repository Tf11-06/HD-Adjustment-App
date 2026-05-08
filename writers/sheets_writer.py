"""
Google Sheets writer for HD Adjustment invoices.
Adapts the existing gspread logic to the Writer ABC.
"""

import os
import re
import gspread
from openpyxl.utils import get_column_letter

import config
from pdf_parser import HEADER_COLS, LI1_COLS, LI_COLS
from .base import Writer

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

_N_INV = len(HEADER_COLS)
_N_LI1 = len(LI1_COLS)
_N_LI  = len(LI_COLS)

_INV_WIDTHS = [14, 16, 14, 14, 14, 14, 14, 18, 10, 22, 8]
_LI1_WIDTHS = [12, 18, 12, 11]
_LI_WIDTHS = [12, 10, 18, 12, 12, 7, 7, 11, 11]


def _rgb(hex_color: str) -> dict:
    hex_color = hex_color.replace("#", "")
    return {
        "red": int(hex_color[0:2], 16) / 255,
        "green": int(hex_color[2:4], 16) / 255,
        "blue": int(hex_color[4:6], 16) / 255,
    }


_HDR1_INV_FMT = {
    "backgroundColor": _rgb("1F2D3D"),
    "horizontalAlignment": "CENTER",
    "verticalAlignment": "MIDDLE",
    "textFormat": {"bold": True, "foregroundColor": _rgb("FFFFFF"), "fontSize": 10},
}
_HDR1_LI1_FMT = {
    "backgroundColor": _rgb("1A4A6E"),
    "horizontalAlignment": "CENTER",
    "verticalAlignment": "MIDDLE",
    "textFormat": {"bold": True, "foregroundColor": _rgb("FFFFFF"), "fontSize": 10},
}
_HDR1_LID_FMT = {
    "backgroundColor": _rgb("5D2510"),
    "horizontalAlignment": "CENTER",
    "verticalAlignment": "MIDDLE",
    "textFormat": {"bold": True, "foregroundColor": _rgb("FFFFFF"), "fontSize": 10},
}
_HDR2_INV_FMT = {
    "backgroundColor": _rgb("344D6E"),
    "horizontalAlignment": "LEFT",
    "verticalAlignment": "MIDDLE",
    "textFormat": {"bold": True, "foregroundColor": _rgb("FFFFFF"), "fontSize": 10},
}
_HDR2_LI1_FMT = {
    "backgroundColor": _rgb("1A4A6E"),
    "horizontalAlignment": "LEFT",
    "verticalAlignment": "MIDDLE",
    "textFormat": {"bold": True, "foregroundColor": _rgb("FFFFFF"), "fontSize": 10},
}
_HDR2_LID_FMT = {
    "backgroundColor": _rgb("5D2510"),
    "horizontalAlignment": "LEFT",
    "verticalAlignment": "MIDDLE",
    "textFormat": {"bold": True, "foregroundColor": _rgb("FFFFFF"), "fontSize": 10},
}
_DATA_INV_FMT = {
    "backgroundColor": _rgb("F9FAFB"),
    "horizontalAlignment": "LEFT",
    "verticalAlignment": "MIDDLE",
    "textFormat": {"fontSize": 10},
}
_DATA_LI1_FMT = {
    "backgroundColor": _rgb("EAF4FB"),
    "horizontalAlignment": "LEFT",
    "verticalAlignment": "MIDDLE",
    "textFormat": {"fontSize": 10},
}
_DATA_LID_FMT = {
    "backgroundColor": _rgb("FDF2E9"),
    "horizontalAlignment": "LEFT",
    "verticalAlignment": "MIDDLE",
    "textFormat": {"fontSize": 10},
}


def _build_header_rows(max_li: int) -> list[list]:
    """Build the 2-row header: [group-label row, field-name row]."""
    row1 = ["Invoice Details"] + [""] * (_N_INV - 1)
    row2 = list(HEADER_COLS)

    # LI1 is the compact credit summary; LI2+ are full debit item groups.
    row1.append("LINE ITEM 1 · Credit")
    row1.extend([""] * (_N_LI1 - 1))
    row2.extend(LI1_COLS)

    for debit in range(max_li):
        row1.append(f"LINE ITEM {debit + 2} · Debit")
        row1.extend([""] * (_N_LI - 1))
        row2.extend(LI_COLS)

    return [row1, row2]


def _a1(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def _a1_range(row1: int, col1: int, row2: int, col2: int) -> str:
    return f"{_a1(row1, col1)}:{_a1(row2, col2)}"


def _li1_start() -> int:
    return _N_INV + 1


def _debit_start(debit_index: int) -> int:
    return _N_INV + _N_LI1 + 1 + debit_index * _N_LI


def _count_li_groups(all_rows: list) -> int:
    """Count LI groups from the group-label row (row index 0)."""
    if not all_rows:
        return 0
    return sum(1 for cell in all_rows[0] if re.match(r'^LINE ITEM \d+', str(cell)))


def _width_to_pixels(width: int) -> int:
    return int(width * 7 + 5)


class SheetsWriter(Writer):
    """Appends invoices to a Google Sheet worksheet."""

    def __init__(self, config: dict):
        self._config = config
        self._worksheet: gspread.Worksheet | None = None
        self._all_rows: list | None = None

    # ── connection ────────────────────────────────────────────────────────

    def connect(self) -> None:
        """Authenticate and open the worksheet. Raises on failure."""
        creds_path = config.resolve_credentials_path(self._config.get("credentials_file", ""))

        if not os.path.exists(creds_path):
            raise FileNotFoundError(
                f"Credentials file not found: {creds_path}. Check the path in Settings."
            )

        try:
            client = gspread.service_account(filename=creds_path, scopes=SCOPES)
        except Exception as e:
            raise ConnectionError(f"Could not authenticate with Google: {e}") from e

        try:
            spreadsheet = client.open_by_key(self._config["sheet_id"])
        except gspread.exceptions.SpreadsheetNotFound:
            raise ValueError("Could not find the Google Sheet. Verify the Sheet ID in Settings.")
        except Exception as e:
            raise ConnectionError(f"Could not connect to Google Sheets: {e}") from e

        name = self._config.get("worksheet_name", "Adjustments")
        try:
            self._worksheet = spreadsheet.worksheet(name)
        except gspread.exceptions.WorksheetNotFound:
            self._worksheet = spreadsheet.add_worksheet(title=name, rows=1000, cols=300)

        self._all_rows = self._worksheet.get_all_values()

    def _require_connected(self):
        if self._worksheet is None:
            raise RuntimeError("SheetsWriter.connect() must be called before use.")

    # ── formatting ───────────────────────────────────────────────────────

    def _safe_merge(self, cell_range: str) -> None:
        try:
            self._worksheet.merge_cells(cell_range)
        except Exception:
            # A reused sheet may already have this merged. Formatting should not
            # block data export if Google rejects a duplicate merge request.
            pass

    def _format_headers(self, max_debit_items: int) -> None:
        self._require_connected()

        li1_start = _li1_start()
        formats = [
            {"range": _a1_range(1, 1, 1, _N_INV), "format": _HDR1_INV_FMT},
            {"range": _a1_range(2, 1, 2, _N_INV), "format": _HDR2_INV_FMT},
            {"range": _a1_range(1, li1_start, 1, li1_start + _N_LI1 - 1), "format": _HDR1_LI1_FMT},
            {"range": _a1_range(2, li1_start, 2, li1_start + _N_LI1 - 1), "format": _HDR2_LI1_FMT},
        ]

        self._safe_merge(_a1_range(1, 1, 1, _N_INV))
        self._safe_merge(_a1_range(1, li1_start, 1, li1_start + _N_LI1 - 1))

        for debit in range(max_debit_items):
            start = _debit_start(debit)
            end = start + _N_LI - 1
            formats.extend([
                {"range": _a1_range(1, start, 1, end), "format": _HDR1_LID_FMT},
                {"range": _a1_range(2, start, 2, end), "format": _HDR2_LID_FMT},
            ])
            self._safe_merge(_a1_range(1, start, 1, end))

        self._worksheet.batch_format(formats)
        self._worksheet.freeze(rows=2)
        self._resize_columns(max_debit_items)

    def _resize_columns(self, max_debit_items: int) -> None:
        spreadsheet = getattr(self._worksheet, "spreadsheet", None)
        if spreadsheet is None:
            try:
                self._worksheet.columns_auto_resize(0, _debit_start(max_debit_items) - 1)
            except Exception:
                pass
            return

        requests = []
        widths = _INV_WIDTHS + _LI1_WIDTHS + (_LI_WIDTHS * max_debit_items)
        for col_index, width in enumerate(widths):
            requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": self._worksheet.id,
                        "dimension": "COLUMNS",
                        "startIndex": col_index,
                        "endIndex": col_index + 1,
                    },
                    "properties": {"pixelSize": _width_to_pixels(width)},
                    "fields": "pixelSize",
                }
            })
        requests.extend([
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": self._worksheet.id,
                        "dimension": "ROWS",
                        "startIndex": 0,
                        "endIndex": 1,
                    },
                    "properties": {"pixelSize": 32},
                    "fields": "pixelSize",
                }
            },
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": self._worksheet.id,
                        "dimension": "ROWS",
                        "startIndex": 1,
                        "endIndex": 2,
                    },
                    "properties": {"pixelSize": 27},
                    "fields": "pixelSize",
                }
            },
        ])
        spreadsheet.batch_update({"requests": requests})

    def _format_data_row(self, row_index: int, max_debit_items: int) -> None:
        self._require_connected()

        li1_start = _li1_start()
        formats = [
            {"range": _a1_range(row_index, 1, row_index, _N_INV), "format": _DATA_INV_FMT},
            {
                "range": _a1_range(row_index, li1_start, row_index, li1_start + _N_LI1 - 1),
                "format": _DATA_LI1_FMT,
            },
        ]
        for debit in range(max_debit_items):
            start = _debit_start(debit)
            formats.append({
                "range": _a1_range(row_index, start, row_index, start + _N_LI - 1),
                "format": _DATA_LID_FMT,
            })
        self._worksheet.batch_format(formats)

    # ── Writer interface ─────────────────────────────────────────────────

    def is_initialized(self) -> bool:
        self._require_connected()
        if not self._all_rows or len(self._all_rows) < 2:
            return False
        return bool(self._all_rows[0]) and str(self._all_rows[1][0]) == HEADER_COLS[0]

    def initialize_headers(self, max_line_items: int = 0) -> None:
        self._require_connected()
        header = _build_header_rows(max_line_items)
        self._worksheet.update(header, "A1")
        self._format_headers(max_line_items)
        self._all_rows = header

    def find_duplicate(self, invoice_num: str) -> bool:
        self._require_connected()
        if not self._all_rows:
            return False
        for row in self._all_rows[2:]:
            if row and row[0] == invoice_num:
                return True
        return False

    def expand_columns_if_needed(self, num_debit_items: int) -> None:
        self._require_connected()
        current = _count_li_groups(self._all_rows)
        needed = num_debit_items + 1  # +1 for LI1 (credit slot always present)
        if needed <= current:
            return
        new_header = _build_header_rows(num_debit_items)
        self._worksheet.update(new_header, "A1")
        self._format_headers(num_debit_items)
        data_rows = self._all_rows[2:] if len(self._all_rows) >= 2 else []
        self._all_rows = new_header + data_rows

    def append_invoice(self, invoice: dict) -> None:
        self._require_connected()

        if not self.is_initialized():
            self.initialize_headers(len(invoice.get('debit_items', [])))
        else:
            self.expand_columns_if_needed(len(invoice.get('debit_items', [])))

        current_li = _count_li_groups(self._all_rows)
        self._format_headers(current_li - 1)

        # Build flat row
        row = self.invoice_header_row(invoice)

        # LI1 — credit summary
        row.extend(self.li1_row(invoice.get('credit_line')))

        # LI2+ — debit items
        for item in invoice.get('debit_items', []):
            row.extend(self.debit_row(item))

        # Pad to full width
        total_li = current_li
        used_li = len(invoice.get('debit_items', [])) + 1  # +1 for LI1
        if used_li < total_li:
            row.extend([''] * (total_li - used_li) * _N_LI)

        next_row = len(self._all_rows) + 1
        self._worksheet.append_rows([row], value_input_option='RAW')
        self._format_data_row(next_row, current_li - 1)
        self._all_rows.append(row)

    # ── convenience ──────────────────────────────────────────────────────

    @property
    def all_rows(self) -> list:
        return self._all_rows or []
