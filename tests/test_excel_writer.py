"""Tests for writers.ExcelWriter."""

import os
import tempfile
import pytest
from openpyxl import load_workbook

from writers.excel_writer import ExcelWriter
from pdf_parser import HEADER_COLS, LI_COLS

# ── fixture invoice dicts ──────────────────────────────────────────────────────

_INVOICE_1 = {
    "invoice_num": "7573", "order_num": "1099173067", "adj_num": "9999",
    "adj_date": "2026-02-24", "inv_date": "2026-01-28", "po_date": "2026-01-26",
    "credit_debit": "C - Credit", "amount": 278.03, "handling": "A - Off Invoice",
    "store": "5089", "vendor_num": "000873237 / 580025", "dept": "28",
    "credit_line": {
        "adj_reason": "24", "sellers_inv": "7573",
        "line_cd": "C - Credit", "item_total": 278.03,
    },
    "debit_items": [
        {"sku": "175525", "vendor_pn": "900690", "adj_reason": "06",
         "sellers_inv": "7573", "line_cd": "D - Debit",
         "qty": 12, "unit": "EA", "unit_price": 3.34, "item_total": 40.08},
        {"sku": "588978", "vendor_pn": "900701", "adj_reason": "06",
         "sellers_inv": "7573", "line_cd": "D - Debit",
         "qty": 48, "unit": "EA", "unit_price": 4.98, "item_total": 239.04},
    ],
    "_file": "invoice_7573.pdf",
}

_INVOICE_2 = {
    "invoice_num": "8888", "order_num": "2222222222", "adj_num": "1",
    "adj_date": "2026-03-01", "inv_date": "2026-02-15", "po_date": "2026-02-10",
    "credit_debit": "D - Debit", "amount": 40.08, "handling": "A - Off Invoice",
    "store": "1001", "vendor_num": "000873237", "dept": "28",
    "credit_line": None,
    "debit_items": [
        {"sku": "175525", "vendor_pn": "900690", "adj_reason": "06",
         "sellers_inv": "8888", "line_cd": "D - Debit",
         "qty": 12, "unit": "EA", "unit_price": 3.34, "item_total": 40.08},
    ],
    "_file": "invoice_8888.pdf",
}


@pytest.fixture
def xlsx_path(tmp_path):
    return str(tmp_path / "test_output.xlsx")


# ── is_initialized ────────────────────────────────────────────────────────────

def test_not_initialized_when_file_missing(xlsx_path):
    w = ExcelWriter(xlsx_path)
    assert not w.is_initialized()


def test_initialized_after_initialize_headers(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(0)
    assert w.is_initialized()


# ── initialize_headers ────────────────────────────────────────────────────────

def test_header_row2_has_invoice_cols(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(0)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for ci, name in enumerate(HEADER_COLS, start=1):
        assert ws.cell(2, ci).value == name


def test_header_row2_has_li_cols(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(2)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    # LI1 starts at col 13
    li1_start = len(HEADER_COLS) + 1
    for ci, name in enumerate(LI_COLS, start=0):
        assert ws.cell(2, li1_start + ci).value == name


def test_header_row1_li1_label(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(1)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    li1_start = len(HEADER_COLS) + 1
    assert "LINE ITEM 1" in str(ws.cell(1, li1_start).value)


def test_freeze_panes_at_a3(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(0)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws.freeze_panes == "A3"


# ── find_duplicate ────────────────────────────────────────────────────────────

def test_find_duplicate_returns_false_on_empty(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(0)
    assert not w.find_duplicate("7573")


def test_find_duplicate_returns_true_after_append(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(2)
    w.append_invoice(_INVOICE_1)
    assert w.find_duplicate("7573")


def test_find_duplicate_returns_false_for_different_invoice(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(2)
    w.append_invoice(_INVOICE_1)
    assert not w.find_duplicate("9999")


# ── append_invoice ────────────────────────────────────────────────────────────

def test_append_invoice_creates_data_row(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(2)
    w.append_invoice(_INVOICE_1)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws.cell(3, 1).value == "7573"


def test_append_invoice_header_fields_correct(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(2)
    w.append_invoice(_INVOICE_1)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws.cell(3, 2).value == "1099173067"   # order_num
    assert ws.cell(3, 7).value == "C - Credit"   # credit_debit
    assert ws.cell(3, 8).value == pytest.approx(278.03)  # amount


def test_rule1_li1_only_has_allowed_fields(xlsx_path):
    """Rule 1: LI1 SKU (col 13), Vendor PN (14), QTY (18), Unit (19), UnitPrice (20) must be blank."""
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(2)
    w.append_invoice(_INVOICE_1)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    li1_start = len(HEADER_COLS) + 1  # col 13
    assert ws.cell(3, li1_start).value in (None, "")      # SKU — blank
    assert ws.cell(3, li1_start + 1).value in (None, "")  # Vendor PN — blank
    assert ws.cell(3, li1_start + 2).value == "24"        # Adj Reason — populated
    assert ws.cell(3, li1_start + 3).value == "7573"      # Sellers Inv — populated
    assert ws.cell(3, li1_start + 4).value == "C - Credit" # Line C/D — populated
    assert ws.cell(3, li1_start + 5).value in (None, "")  # QTY — blank
    assert ws.cell(3, li1_start + 6).value in (None, "")  # Unit — blank
    assert ws.cell(3, li1_start + 7).value in (None, "")  # Unit Price — blank
    assert ws.cell(3, li1_start + 8).value == pytest.approx(278.03)  # Item Total — populated


def test_rule3_compact_format_li1_all_blank(xlsx_path):
    """Rule 3: no credit_line → all LI1 cells blank."""
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(1)
    w.append_invoice(_INVOICE_2)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    li1_start = len(HEADER_COLS) + 1
    for offset in range(9):
        assert ws.cell(3, li1_start + offset).value in (None, "")


def test_debit_items_start_at_li2(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(2)
    w.append_invoice(_INVOICE_1)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    li2_start = len(HEADER_COLS) + 1 + len(LI_COLS)  # col 22
    assert ws.cell(3, li2_start).value == "175525"      # SKU of debit item 1
    assert ws.cell(3, li2_start + 5).value == 12        # QTY


def test_append_two_invoices_two_rows(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(2)
    w.append_invoice(_INVOICE_1)
    w.append_invoice(_INVOICE_2)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws.cell(3, 1).value == "7573"
    assert ws.cell(4, 1).value == "8888"


# ── expand_columns_if_needed ──────────────────────────────────────────────────

def test_expand_adds_new_li_column_group(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(1)   # starts with space for 1 debit item
    w.expand_columns_if_needed(3)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    # col for LI4 (index 3): 12 + 9*3 + 1 = 40
    li4_start = len(HEADER_COLS) + 1 + 3 * len(LI_COLS)
    assert ws.cell(2, li4_start).value == LI_COLS[0]   # "SKU"


def test_expand_does_not_shrink(xlsx_path):
    w = ExcelWriter(xlsx_path)
    w.initialize_headers(5)
    w.expand_columns_if_needed(1)  # fewer than existing — no change
    wb = load_workbook(xlsx_path)
    ws = wb.active
    # LI6 header should still exist
    li6_start = len(HEADER_COLS) + 1 + 5 * len(LI_COLS)
    assert ws.cell(2, li6_start).value == LI_COLS[0]
